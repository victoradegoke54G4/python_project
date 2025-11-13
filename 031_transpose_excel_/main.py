
import openpyxl


input_ssh = openpyxl.load_workbook('031_transpose_excel_\sample_data.xlsx')
input_sheet = input_ssh.active
output_wb = openpyxl.Workbook()
output_sheet = output_wb.active
col_1 = []
col_2 = []


for row in range(1, input_sheet.max_row + 1):
    for col in range(1, input_sheet.max_column + 1):
        if col == 1:
            col_val_1 = input_sheet.cell(row=row, column=col).value
            col_1.append(col_val_1)
        else:
            col_val_2 = input_sheet.cell(row=row, column=col).value
            col_2.append(col_val_2)


col_count = 1
col_count_2 = 1
for item in col_1:
    output_sheet.cell(row=1, column=col_count).value = item
    col_count += 1
for item in col_2:
    output_sheet.cell(row=2, column=col_count_2).value = item
    col_count_2 += 1
    
output_wb.save('031_transpose_excel_\output.xlsx')
